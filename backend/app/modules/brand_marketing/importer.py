from contextlib import closing
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlmodel import Session

from app.modules.brand_marketing.constants import (
    URBAN_POPULATION_INDICATOR_CODE,
    RegionalIndicatorCode,
)
from app.modules.brand_marketing.models import ProvinceAnnualIndicator

DATA_DIR = Path(__file__).resolve().parents[4] / "data"

INDICATOR_FILENAMES = {
    RegionalIndicatorCode.RESIDENT_POPULATION.value: "年末常住人口 (万人).xlsx",
    URBAN_POPULATION_INDICATOR_CODE: "城镇人口 (万人).xlsx",
    RegionalIndicatorCode.PER_CAPITA_GDP.value: "人均地区生产总值 (元人).xlsx",
    RegionalIndicatorCode.DISPOSABLE_INCOME.value: "全体居民人均可支配收入 (元).xlsx",
    RegionalIndicatorCode.CONSUMPTION_EXPENDITURE.value: "全体居民人均消费支出 (元).xlsx",
    RegionalIndicatorCode.RETAIL_SALES.value: "社会消费品零售总额 (亿元).xlsx",
}

PROVINCE_CODES = {
    "北京市": "110000",
    "天津市": "120000",
    "河北省": "130000",
    "山西省": "140000",
    "内蒙古自治区": "150000",
    "辽宁省": "210000",
    "吉林省": "220000",
    "黑龙江省": "230000",
    "上海市": "310000",
    "江苏省": "320000",
    "浙江省": "330000",
    "安徽省": "340000",
    "福建省": "350000",
    "江西省": "360000",
    "山东省": "370000",
    "河南省": "410000",
    "湖北省": "420000",
    "湖南省": "430000",
    "广东省": "440000",
    "广西壮族自治区": "450000",
    "海南省": "460000",
    "重庆市": "500000",
    "四川省": "510000",
    "贵州省": "520000",
    "云南省": "530000",
    "西藏自治区": "540000",
    "陕西省": "610000",
    "甘肃省": "620000",
    "青海省": "630000",
    "宁夏回族自治区": "640000",
    "新疆维吾尔自治区": "650000",
}


def import_regional_data(*, session: Session) -> None:
    """导入国家统计局分省年度数据。"""
    records: list[ProvinceAnnualIndicator] = []

    for indicator_code, filename in INDICATOR_FILENAMES.items():
        records.extend(_read_workbook(indicator_code, filename))

    session.exec(delete(ProvinceAnnualIndicator))
    session.add_all(records)
    session.commit()


def _read_workbook(
    indicator_code: str,
    filename: str,
) -> list[ProvinceAnnualIndicator]:
    with closing(load_workbook(DATA_DIR / filename, read_only=True)) as workbook:
        sheet = workbook.worksheets[0]

        if sheet.cell(4, 1).value != "地区":
            raise ValueError(f"{filename} 第 4 行不是数据表头")

        years: list[int] = []

        for column in range(2, sheet.max_column + 1):
            year = int(str(sheet.cell(4, column).value).removesuffix("年"))
            years.append(year)

        if not years or len(years) != len(set(years)):
            raise ValueError(f"{filename} 年份为空或重复")

        records: list[ProvinceAnnualIndicator] = []
        province_names: set[str] = set()

        for row in sheet.iter_rows(
            min_row=5,
            max_row=4 + len(PROVINCE_CODES),
            max_col=len(years) + 1,
            values_only=True,
        ):
            province_name = row[0]

            if not isinstance(province_name, str):
                raise ValueError(f"{filename} 包含无效地区")

            if province_name not in PROVINCE_CODES or province_name in province_names:
                raise ValueError(f"{filename} 包含未知或重复地区：{province_name}")

            province_names.add(province_name)

            for column_index, year in enumerate(years, start=1):
                value = row[column_index]

                if value is None or value == "":
                    continue

                records.append(
                    ProvinceAnnualIndicator(
                        indicator_code=indicator_code,
                        province_code=PROVINCE_CODES[province_name],
                        province_name=province_name,
                        year=year,
                        value=Decimal(str(value)),
                    )
                )

        if province_names != PROVINCE_CODES.keys():
            raise ValueError(f"{filename} 地区数据不完整")

        return records
